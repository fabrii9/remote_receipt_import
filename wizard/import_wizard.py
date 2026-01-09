# -*- coding: utf-8 -*-
import base64
import io
import re
import xmlrpc.client
import time
import random
import logging
from datetime import datetime
from collections import defaultdict

from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    import openpyxl
except Exception:
    openpyxl = None

import csv


class RemotePaymentImport(models.Model):
    _name = "remote.payment.import"
    _description = "Importar pagos y crear recibos en Odoo 18"
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = "create_date desc"
    _rec_name = "filename"

    upload = fields.Binary(string="Archivo (XLSX o CSV)", attachment=True)
    filename = fields.Char(string="Nombre de archivo", tracking=True)
    
    # Campos de progreso y estado
    state = fields.Selection([
        ('draft', 'Borrador'),
        ('processing', 'Procesando...'),
        ('done', 'Completado'),
    ], string="Estado", default='draft', readonly=True, tracking=True)
    progress_message = fields.Text(string="Progreso", readonly=True)
    total_rows = fields.Integer(string="Total filas", readonly=True)
    processed_rows = fields.Integer(string="Filas procesadas", readonly=True)
    
    # Relaciones con logs
    batch_id = fields.Many2one('remote.payment.import.log', string="Batch Log", readonly=True)
    checkpoint_id = fields.Many2one('payment.import.checkpoint', string="Checkpoint", readonly=True)
    
    # Acción para ver dashboard
    def action_view_dashboard(self):
        self.ensure_one()
        if not self.checkpoint_id:
            raise UserError(_("No hay checkpoint asociado a esta importación"))
        return {
            'type': 'ir.actions.act_window',
            'name': _('Dashboard de Progreso'),
            'res_model': 'payment.import.checkpoint',
            'res_id': self.checkpoint_id.id,
            'view_mode': 'form',
            'target': 'current',
        }
    
    # Acción para ver cola
    def action_view_queue(self):
        self.ensure_one()
        if not self.batch_id:
            raise UserError(_("No hay batch asociado a esta importación"))
        return {
            'type': 'ir.actions.act_window',
            'name': _('Cola de Procesamiento'),
            'res_model': 'payment.import.queue.line',
            'view_mode': 'tree,form',
            'domain': [('batch_id', '=', self.batch_id.id)],
            'context': {'default_batch_id': self.batch_id.id},
            'target': 'current',
        }

    # -------------------------
    # Helpers de configuración
    # -------------------------
    def _read_settings(self):
        """Lee la configuración desde el modelo remote.receipt.settings"""
        settings = self.env['remote.receipt.settings'].sudo().search([], limit=1, order='id desc')
        
        if not settings:
            # Fallback: intentar leer desde ir.config_parameter (retrocompatibilidad)
            ICP = self.env['ir.config_parameter'].sudo()
            url = ICP.get_param("remote_receipt_import.remote_o18_url")
            db = ICP.get_param("remote_receipt_import.remote_o18_db")
            user = ICP.get_param("remote_receipt_import.remote_o18_user")
            pwd = ICP.get_param("remote_receipt_import.remote_o18_password")
            journal_id = int(ICP.get_param("remote_receipt_import.remote_payment_journal_id") or 0)
            pm_line_id = int(ICP.get_param("remote_receipt_import.remote_payment_method_line_id") or 0)
            tol = float(ICP.get_param("remote_receipt_import.amount_tolerance") or 0.01)
            
            if not url or not db or not user or not pwd or not journal_id:
                raise UserError(_("Debes configurar la conexión primero. Ve a Contabilidad → Importación Remota → Configuración."))
            return url, db, user, pwd, journal_id, pm_line_id, tol
        
        # Leer desde el modelo
        if not settings.remote_o18_url or not settings.remote_o18_db or not settings.remote_o18_user or not settings.remote_o18_password or not settings.remote_payment_journal_id:
            raise UserError(_("La configuración está incompleta. Ve a Contabilidad → Importación Remota → Configuración."))
        
        return (
            settings.remote_o18_url,
            settings.remote_o18_db,
            settings.remote_o18_user,
            settings.remote_o18_password,
            settings.remote_payment_journal_id,  # Ya es int, no necesita .id
            settings.remote_payment_method_line_id or 0,  # Ya es int
            settings.amount_tolerance
        )

    def _xmlrpc_env(self, url, db, user, pwd):
        common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common")
        uid = common.authenticate(db, user, pwd, {})
        if not uid:
            raise UserError(_("No se pudo autenticar en Odoo 18 con las credenciales provistas."))
        objects = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object")
        return uid, objects


    # -------------------------
    # XML-RPC con retry (anti 429 / rate limit)
    # -------------------------
    def _execute_kw_with_retry(
        self,
        objects,
        db,
        uid,
        pwd,
        model,
        method,
        args,
        kwargs=None,
        max_retries=6,
        base_backoff=1.5,
        max_sleep=20.0,
    ):
        """Wrapper centralizado para execute_kw con reintentos ante HTTP 429."""
        kwargs = kwargs or {}
        attempt = 0
        while True:
            try:
                return objects.execute_kw(db, uid, pwd, model, method, args, kwargs)
            except xmlrpc.client.ProtocolError as e:
                # 429 Too Many Requests
                if getattr(e, "errcode", None) == 429 and attempt < max_retries:
                    attempt += 1
                    delay = min(max_sleep, base_backoff * attempt) + random.uniform(0, 0.4)
                    _logger = logging.getLogger(__name__)
                    _logger.warning(
                        "XML-RPC 429 en %s.%s intento=%s/%s; durmiendo %.2fs",
                        model, method, attempt, max_retries, delay
                    )
                    time.sleep(delay)
                    continue
                raise


    # -------------------------
    # Idempotencia (evitar duplicados)
    # -------------------------
    def _make_idempotency_key(
        self,
        journal_id: int,
        company_id: int,
        partner_id: int,
        amount: float,
        date_str: str,
        memo_raw: str,
    ) -> str:
        """Genera una clave estable para identificar un recibo.

        Motivo: con rate-limit (HTTP 429) y/o ejecuciones concurrentes, el wizard puede
        intentar crear el mismo pago más de una vez. Esta clave se graba en `ref` y se
        usa para buscar pagos ya creados.
        """
        memo = (memo_raw or "").strip()
        memo_norm = re.sub(r"\s+", " ", memo)[:120]
        return f"RRI|j{int(journal_id)}|c{int(company_id)}|p{int(partner_id)}|a{amount:.2f}|d{date_str}|m{memo_norm}"[:250]


    def _find_existing_payment(self, objects, db, uid, pwd, ctx, idem_key: str):
        """Busca un payment existente por ref=idempotency_key."""
        recs = self._execute_kw_with_retry(
            objects,
            db,
            uid,
            pwd,
            "account.payment",
            "search_read",
            [[("ref", "=", idem_key)]],
            {"fields": ["id", "state"], "limit": 1, "context": ctx},
        )
        return recs[0] if recs else None

    def _batch_search_partners(self, objects, db, uid, pwd, ctx, cuit_variants_list, journal_company_id):
        """Busca múltiples partners en una sola llamada XML-RPC.
        
        Args:
            cuit_variants_list: lista de tuplas [(cuit_norm, [variants]), ...]
            journal_company_id: ID de la compañía del diario para priorizar partners
        
        Returns:
            dict: {cuit_normalizado: partner_data o None}
        """
        # Aplanar todas las variantes y crear dominio OR masivo
        all_variants = []
        variant_to_original = {}  # mapeo de variante a CUIT normalizado original
        
        for cuit_norm, variants in cuit_variants_list:
            for v in variants:
                all_variants.append(v)
                variant_to_original[v] = cuit_norm
        
        if not all_variants:
            return {}
        
        # Crear dominio OR para todas las variantes
        clauses = []
        for v in all_variants:
            clauses.extend([
                ("vat", "=", v),
                ("ref", "=", v),
                ("commercial_partner_id.vat", "=", v),
            ])
        domain = ["|"] * (len(clauses) - 1) + clauses if clauses else [("id", "=", 0)]
        
        # Búsqueda batch
        partner_ids = self._execute_kw_with_retry(
            objects, db, uid, pwd, "res.partner", "search",
            [domain],
            {"limit": 200, "context": ctx}
        )
        
        # Fallback ILIKE si no encontró por igualdad
        if not partner_ids:
            clauses_ilike = []
            for v in all_variants:
                clauses_ilike.extend([
                    ("vat", "ilike", v),
                    ("ref", "ilike", v),
                    ("commercial_partner_id.vat", "ilike", v),
                ])
            domain_ilike = ["|"] * (len(clauses_ilike) - 1) + clauses_ilike if clauses_ilike else [("id", "=", 0)]
            partner_ids = self._execute_kw_with_retry(
                objects, db, uid, pwd, "res.partner", "search",
                [domain_ilike],
                {"limit": 200, "context": ctx}
            )
        
        if not partner_ids:
            return {cuit: None for cuit, _ in cuit_variants_list}
        
        # Leer todos los partners encontrados
        partners_data = self._execute_kw_with_retry(
            objects, db, uid, pwd, "res.partner", "read",
            [partner_ids, ["name", "company_id", "vat", "ref"]],
            {"context": ctx}
        )
        
        # Helper para extraer ID de campo many2one
        def _m2o_id(val):
            if isinstance(val, (list, tuple)) and val:
                return val[0]
            if isinstance(val, int):
                return val
            return False
        
        # Agrupar partners por CUIT normalizado
        cuit_to_partners = {}
        for p in partners_data:
            pvat = self._normalize_cuit(p.get("vat"))
            pref = self._normalize_cuit(p.get("ref"))
            
            for variant in all_variants:
                vnorm = self._normalize_cuit(variant)
                if vnorm and (vnorm == pvat or vnorm == pref):
                    original_cuit = variant_to_original.get(variant)
                    if original_cuit:
                        if original_cuit not in cuit_to_partners:
                            cuit_to_partners[original_cuit] = []
                        cuit_to_partners[original_cuit].append(p)
                    break
        
        # Elegir el partner correcto por compañía (mismo algoritmo que antes)
        result = {}
        for cuit_norm, _ in cuit_variants_list:
            partners_list = cuit_to_partners.get(cuit_norm, [])
            if not partners_list:
                result[cuit_norm] = None
                continue
            
            # Elegir partner por compañía del diario (o sin compañía)
            chosen = None
            fallback_none_company = None
            for p in partners_list:
                cid = _m2o_id(p.get("company_id"))
                if cid == journal_company_id:
                    chosen = p
                    break
                if not cid and not fallback_none_company:
                    fallback_none_company = p
            
            if not chosen:
                chosen = fallback_none_company or partners_list[0]
            
            result[cuit_norm] = chosen
        
        return result


    # -------------------------
    # Normalizadores / parsing
    # -------------------------
    def _normalize_cuit(self, raw):
        """Devuelve solo dígitos; maneja números y notación científica de Excel."""
        if raw is None:
            return ""
        if isinstance(raw, (int, float)):
            try:
                as_int = int(round(float(raw)))
                return str(as_int)
            except Exception:
                return re.sub(r"\D", "", str(raw))
        s = str(raw).strip()
        try:
            # Notación científica como "1.23254E+11"
            if 'e' in s.lower():
                num = float(s.replace(",", "."))
                return str(int(round(num)))
        except Exception:
            pass
        return re.sub(r"\D", "", s)

    def _vat_variants(self, cuit_raw, cuit_digits):
        """
        Variantes comunes para buscar el CUIT/DNI:
        - valor original (tal cual viene del Excel)
        - solo dígitos
        - CUIT con guiones: XX-XXXXXXXX-X
        - DNI con puntos: XX.XXX.XXX (si tiene 8 dígitos)
        """
        variants = set()
        
        # Agregar valor original tal cual está en el Excel
        if cuit_raw:
            original = str(cuit_raw).strip()
            if original:
                variants.add(original)
        
        # Agregar variantes normalizadas
        s = re.sub(r"\D", "", str(cuit_digits or ""))
        if not s:
            return list(variants) if variants else []
        
        variants.add(s)
        if len(s) == 11:  # CUIT
            variants.add(f"{s[:2]}-{s[2:10]}-{s[10:]}")
        if len(s) == 8:   # DNI
            variants.add(f"{s[:2]}.{s[2:5]}.{s[5:]}")
        return list(variants)

    def _parse_amount(self, raw):
        if raw is None:
            return 0.0
        if isinstance(raw, (int, float)):
            return float(raw)
        s = str(raw).strip()
        if "," in s and "." not in s:
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(" ", "").replace(",", "")
        try:
            return float(s)
        except Exception:
            return 0.0

    def _parse_date(self, raw):
        if not raw:
            return fields.Date.context_today(self)
        if isinstance(raw, datetime):
            return raw.date()
        if isinstance(raw, (float, int)) and openpyxl:
            try:
                from openpyxl.utils.datetime import from_excel
                return from_excel(raw).date()
            except Exception:
                pass
        s = str(raw).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                continue
        return fields.Date.context_today(self)

    # -------------------------
    # Lectura del archivo
    # -------------------------
    def _read_rows(self, content, filename):
        """
        Devuelve lista de dicts con llaves:
        fecha_pago, tipo_operacion (CUIT/DNI), operacion_relacionada (para memo), importe

        ⚠️ Mapeo correcto:
        - CUIT/DNI = **Tipo de Operación**
        - MEMO     = **Operación Relacionada**
        """
        name = (filename or "").lower()
        rows = []
        if name.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
            if not openpyxl:
                raise UserError(_("Falta dependencia 'openpyxl' para leer archivos .xlsx"))
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

            def find_col(name_part):
                for idx, h in enumerate(headers):
                    if name_part.lower() in h.lower():
                        return idx
                return None

            c_fecha = find_col("de Pago") or find_col("fecha")
            c_tipo = find_col("Tipo de Operación") or find_col("tipo")
            c_rel = find_col("Operación Relacionada") or find_col("relacionada")
            c_imp = find_col("Importe") or find_col("mporte")

            if c_tipo is None:
                raise UserError(_("No se encontró la columna 'Tipo de Operación' (CUIT/DNI)."))
            if c_imp is None:
                raise UserError(_("No se encontró la columna 'Importe'."))

            for r in ws.iter_rows(min_row=2):
                tipo_val = r[c_tipo].value if c_tipo is not None else None   # CUIT/DNI
                rel_val = r[c_rel].value if c_rel is not None else None     # MEMO
                vals = {
                    "fecha_pago": self._parse_date(r[c_fecha].value) if c_fecha is not None else fields.Date.context_today(self),
                    "tipo_operacion": (tipo_val or ""),             # crudo para mostrar
                    "operacion_relacionada": rel_val,               # crudo para memo
                    "importe": self._parse_amount(r[c_imp].value),
                }
                if (not str(vals["tipo_operacion"]).strip()) and (not vals["importe"]):
                    continue
                rows.append(vals)
        else:
            text = io.StringIO(content.decode("utf-8", errors="ignore"))
            reader = csv.DictReader(text)

            def pick(d, *keys):
                for k in keys:
                    if k in d:
                        return d[k]
                    for dk in d.keys():
                        if k.lower() == dk.lower():
                            return d[dk]
                return None

            for d in reader:
                tipo_val = pick(d, "Tipo de Operación", "Tipo", "Operacion", "Operación")  # CUIT/DNI
                rel_val = pick(d, "Operación Relacionada", "Operacion Relacionada")        # MEMO
                vals = {
                    "fecha_pago": self._parse_date(pick(d, "de Pago", "Fecha de Pago", "Fecha")),
                    "tipo_operacion": (tipo_val or ""),
                    "operacion_relacionada": rel_val,
                    "importe": self._parse_amount(pick(d, "Importe", "Monto", "Total")),
                }
                if (not str(vals["tipo_operacion"]).strip()) and (not vals["importe"]):
                    continue
                rows.append(vals)
        return rows

    # -------------------------
    # Proceso principal (ARQUITECTURA ROBUSTA: Solo crea cola)
    # -------------------------
    def action_process(self):
        """
        Fase 1: Ingesta - Solo valida y crea registros en cola.
        NO procesa, NO llama al remoto.
        El procesamiento lo hace el job asíncrono.
        """
        self.ensure_one()
        if not self.upload or not self.filename:
            raise UserError(_("Debes subir un archivo antes de procesar."))

        _logger = logging.getLogger(__name__)
        
        # Leer y validar archivo
        content = base64.b64decode(self.upload)
        rows = self._read_rows(content, self.filename or "")
        
        if not rows:
            raise UserError(_("El archivo no contiene filas válidas para procesar."))
        
        _logger.info(f"📥 Ingesta: {len(rows)} filas del archivo {self.filename}")
        
        # Crear log/batch
        log = self.env["remote.payment.import.log"].sudo().create({
            "file_name": self.filename or "archivo",
        })
        
        # Crear checkpoint
        checkpoint = self.env["payment.import.checkpoint"].sudo().create({
            "batch_id": log.id,
            "total_rows": len(rows),
            "state": "running",
        })
        
        # Crear registros en cola (batch creation para performance)
        queue_vals = []
        for idx, row in enumerate(rows, start=1):
            import json
            queue_vals.append({
                "batch_id": log.id,
                "row_number": idx,
                "fecha_pago": row.get("fecha_pago"),
                "tipo_operacion": str(row.get("tipo_operacion") or ""),
                "operacion_relacionada": str(row.get("operacion_relacionada") or ""),
                "importe": float(row.get("importe") or 0.0),
                "row_data": json.dumps(row, default=str),  # Serializar para flexibilidad
                "state": "pending",
                "priority": 10,  # Prioridad normal
            })
        
        # Crear todas las líneas en batch
        _logger.info(f"⏳ Creando {len(queue_vals)} registros en cola...")
        self.env["payment.import.queue.line"].sudo().create(queue_vals)
        self.env.cr.commit()
        
        _logger.info(f"✅ Cola creada exitosamente: {len(queue_vals)} registros")
        
        # Actualizar wizard
        self.write({
            'state': 'done',
            'total_rows': len(rows),
            'processed_rows': 0,
            'progress_message': f'✓ Archivo cargado: {len(rows)} pagos en cola\n⏳ El procesamiento comenzará en background\n📊 Revisá el Dashboard de Progreso'
        })
        self.env.cr.commit()
        
        # Notificar al usuario
        try:
            self.env['bus.bus']._sendone(self.env.user.partner_id, 'simple_notification', {
                'type': 'success',
                'title': 'Archivo en Cola',
                'message': f'✓ {len(rows)} pagos listos para procesar\nEl procesamiento se hará en segundo plano de forma segura.',
                'sticky': True,
            })
        except Exception:
            pass
        
        # Encolar el primer job de procesamiento (si queue_job está instalado)
        try:
            if hasattr(self.env["payment.import.queue.line"], "with_delay"):
                # queue_job está disponible
                self.env["payment.import.queue.line"].with_delay(
                    priority=5,
                    description=f"Procesar pagos: {self.filename}"
                ).process_queue_batch(batch_id=log.id, checkpoint_id=checkpoint.id)
                _logger.info(f"🚀 Job encolado para procesar batch {log.id}")
            else:
                # Fallback: será procesado por cron
                _logger.info(f"⏰ queue_job no disponible, será procesado por cron")
        except Exception as e:
            _logger.warning(f"⚠️ No se pudo encolar job: {e}. Será procesado por cron.")
        
        # Retornar a vista de progreso
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'payment.import.checkpoint',
            'res_id': checkpoint.id,
            'view_mode': 'form',
            'target': 'current',
            'context': {'batch_id': log.id}
        }
    
    def action_view_logs(self):
        """Abrir los logs de la importación."""
        self.ensure_one()
        log_id = self.env.context.get('active_log_id')
        if not log_id:
            # Buscar el log más reciente con el mismo nombre de archivo
            log = self.env['remote.payment.import.log'].sudo().search(
                [('file_name', '=', self.filename or 'archivo')],
                order='create_date desc',
                limit=1
            )
            log_id = log.id if log else False
        
        if not log_id:
            raise UserError(_('No se encontró el log de importación.'))
        
        action = self.env.ref('remote_receipt_import.action_remote_payment_import_log').sudo().read()[0]
        action['domain'] = [('id', '=', log_id)]
        action['target'] = 'current'
        return action
